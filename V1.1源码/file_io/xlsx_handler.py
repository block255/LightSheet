"""XLSX 文件读写。使用 openpyxl。

V1.1 起支持多 sheet：
- load_all / write_all 读写整个工作簿（全部 sheet，按标签顺序）
- load / write 保留单 sheet 语义（第一个 sheet），兼容旧调用方
"""
from openpyxl import Workbook, load_workbook

_INVALID_NAME_CHARS = '\\/?*[]:'


def load(file_path: str) -> list[list[str]]:
    """读取 XLSX 文件第一个 sheet，返回二维列表（兼容旧调用方）。"""
    sheets = load_all(file_path)
    return sheets[0][1] if sheets else []


def load_all(file_path: str) -> list[tuple[str, list[list[str]]]]:
    """读取 XLSX 全部 sheet，返回 [(sheet名, 二维列表), ...]（按标签顺序）。"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    result: list[tuple[str, list[list[str]]]] = []
    try:
        for ws in wb.worksheets:
            rows: list[list[str]] = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    val = cell.value
                    row_data.append(str(val) if val is not None else '')
                rows.append(row_data)
            result.append((ws.title, rows))
    finally:
        wb.close()
    return result


def write(file_path: str, data: list[list[str]]) -> None:
    """写入 XLSX 文件（单 sheet，兼容旧调用方）。"""
    write_all(file_path, [('Sheet1', data)])


def write_all(file_path: str, sheets: list[tuple[str, list[list[str]]]],
              formulas: dict = None) -> None:
    """写入 XLSX 文件，包含全部 sheet（按列表顺序）。

    sheet 名自动清理：去除非法字符、限制 31 字符、重名追加序号。
    formulas（互译 P2 可选）：{sheet名: {(r0, c0): 公式文本}}，
    写入时对应格存公式（<f>，给 Excel 看）；无 formulas 行为同旧版。
    """
    wb = Workbook()
    used_names: set[str] = set()
    first = True
    for name, data in sheets:
        title = _unique_sheet_name(name, used_names)
        if first:
            ws = wb.active
            ws.title = title
            first = False
        else:
            ws = wb.create_sheet(title=title)
        for row_data in data:
            ws.append(row_data)
        # 公式写入（覆盖为公式文本，openpyxl 自动识别 '=' 前缀）
        if formulas:
            fm = formulas.get(name)
            if fm:
                for (r, c), f in fm.items():
                    ws.cell(row=r + 1, column=c + 1).value = f
    if first:  # sheets 为空：保留默认空 sheet
        wb.active.title = 'Sheet1'
    wb.save(file_path)
    wb.close()


def _unique_sheet_name(name: str, used: set[str]) -> str:
    """生成合法且不重名的 sheet 名。"""
    base = ''.join(c if c not in _INVALID_NAME_CHARS else '_' for c in str(name))
    if not base:
        base = 'Sheet1'
    base = base[:31]
    title = base
    n = 2
    while title in used:
        suffix = str(n)
        title = f'{base[:31 - len(suffix)]}{suffix}'
        n += 1
    used.add(title)
    return title
