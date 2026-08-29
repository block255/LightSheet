"""公式互译层（Formula Translate）— xlsx 公式格 ↔ 区域引用。

P1 职责（本期）：
  1. scan_workbook：扫描 xlsx 全部 sheet 的公式格（<f> 元素），
     解析每个公式引用的格子区域 → 供 ExtStore.formula_cells 存储；
  2. 引用解析（parse_formula_refs）：公式文本 → [(sheet, (r1,c1,r2,c2)), ...]，
     支持 普通区域/单格/跨sheet/绝对引用/整列整行/共享公式（openpyxl 已展开）。

P2/P3（远期）：脚本 params ↔ 公式文本双向翻译、公式求值引擎。
"""
import re
from typing import Optional

from openpyxl import load_workbook

# ------------------------------------------------------------------
# A1 风格坐标解析
# ------------------------------------------------------------------

_COL_LETTER_RE = re.compile(r'[A-Za-z]+')
_CELL_RE = re.compile(r'([A-Za-z]+)(\d+)')


def col_letter_to_index(letters: str) -> int:
    """列字母 → 0-based 索引（A→0, Z→25, AA→26）。"""
    idx = 0
    for ch in letters.upper():
        idx = idx * 26 + (ord(ch) - 64)
    return idx - 1


def index_to_col_letter(idx: int) -> str:
    """0-based 列索引 → 字母。"""
    s = ''
    n = idx + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord('A') + r) + s
    return s


def parse_cell_ref(ref: str) -> Optional[tuple[int, int]]:
    """'A1' → (row0, col0)；非法返回 None。"""
    m = _CELL_RE.fullmatch(ref.strip())
    if not m:
        return None
    return int(m.group(2)) - 1, col_letter_to_index(m.group(1))


def parse_range_ref(ref: str) -> Optional[tuple[int, int, int, int]]:
    """'A1:B3' → (r1, c1, r2, c2)；单格 'A1' → (r, c, r, c)。"""
    ref = ref.strip()
    if ':' in ref:
        a, b = ref.split(':')
        p1, p2 = parse_cell_ref(a), parse_cell_ref(b)
        if not p1 or not p2:
            return None
        r1, c1 = p1
        r2, c2 = p2
        return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)
    p = parse_cell_ref(ref)
    return (p[0], p[1], p[0], p[1]) if p else None


# ------------------------------------------------------------------
# 公式文本 → 引用区域
# ------------------------------------------------------------------

def parse_formula_refs(formula: str) -> list[tuple[Optional[str], tuple]]:
    """公式文本 → [(sheet名|None, (r1,c1,r2,c2)), ...]。

    支持：A1、A1:B3、数据表!A1、$A$1（去 $）、整列 A:A、整行 1:1。
    跨 sheet 表名：公式中 `!` 前的标识符。
    """
    refs: list[tuple[Optional[str], tuple]] = []
    f = formula.replace('$', '')
    pat = re.compile(
        r"(?<![A-Za-z0-9_])(?:([^!+\-*/%^&<>=(),\s]+)!)?"
        r"((?:[A-Za-z]+:\s*[A-Za-z]+)|(?:[A-Za-z]+\d+:\s*[A-Za-z]+\d+)"
        r"|(?:[A-Za-z]+\d+)|(?:\d+:\d+))"
    )
    for m in pat.finditer(f):
        sheet = m.group(1)
        body = m.group(2).replace(' ', '')
        if ':' in body:
            left, right = body.split(':')
            if left.isalpha() and right.isalpha():
                # 整列 A:A
                c = col_letter_to_index(left)
                refs.append((sheet, (0, c, 10**6, c)))
                continue
            if left.isdigit() and right.isdigit():
                # 整行 1:1
                refs.append((sheet, (int(left) - 1, 0, int(right) - 1, 10**6)))
                continue
            rng = parse_range_ref(body)
            if rng:
                refs.append((sheet, rng))
        else:
            p = parse_cell_ref(body)
            if p:
                refs.append((sheet, (p[0], p[1], p[0], p[1])))
    return refs


# ------------------------------------------------------------------
# 工作簿公式格扫描
# ------------------------------------------------------------------

def scan_workbook(file_path: str) -> list[dict]:
    """扫描 xlsx 全部公式格。

    返回 [{sheet, cell:[r,c], formula, refs:[{'sheet':来源表|None,'range':[...]}...]}, ...]
    仅收集有公式的单元格；openpyxl 读公式时共享公式已自动展开。
    失败（非 xlsx/损坏）返回 []。
    """
    try:
        wb = load_workbook(file_path, data_only=False, read_only=True)
    except Exception:
        return []
    try:
        out = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith('='):
                        refs = []
                        for _sheet, rng in parse_formula_refs(v):
                            refs.append({'sheet': _sheet, 'range': list(rng)})
                        out.append({
                            'sheet': ws.title,
                            'cell': [cell.row - 1, cell.column - 1],
                            'formula': v,
                            'refs': refs,
                        })
        return out
    finally:
        wb.close()


def formula_cell_map(formula_cells: list[dict]) -> dict:
    """formula_cells → {(sheet, r, c): formula} 查找表（动态引擎避开用）。

    key 含 sheet 名：跨 sheet 时同一坐标 (r,c) 在不同表不算冲突。
    """
    return {(fc.get('sheet', ''), fc['cell'][0], fc['cell'][1]): fc['formula']
            for fc in formula_cells}


def workbook_fingerprint(file_path: str) -> str:
    """xlsx 身份指纹：排序后的 sheet 名列表哈希（轻量，只读 sheet 名不读数据）。

    用途（同名文件自动区分）：同名同路径文件被替换/内容大变时 sheet 结构
    变化 → 指纹变化 → 扩展区检测到文件身份改变 → 触发对账重建。
    失败（非 xlsx/损坏）返回空串。
    """
    import hashlib
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        names = sorted(ws.title for ws in wb.worksheets)
        wb.close()
    except Exception:
        return ''
    return hashlib.md5(','.join(names).encode('utf-8')).hexdigest()


# ----------------------------------------------------------------------
# 写方向翻译（互译 P2）：replay_cfg + 脚本名 → 公式模板
# ----------------------------------------------------------------------
# 模板含占位：{r} = 行号（相对输出行）、{c} = 列字母（相对输出列）。
# 展开用 formula_engine.expand_template（P3 引擎已提供）。
# 不可译 → 返回 None（干净丢失，条目保持 script 形态，用我们机制）。

_SCRIPT_AGG_FUNC = {
    '平均': 'AVERAGE', '求和': 'SUM', '最大': 'MAX', '最小': 'MIN',
    '方差': 'VAR', '标准': 'STDEV',
}
_SCRIPT_BIN_OP = {'加法': '+', '减法': '-', '乘法': '*', '除法': '/'}
# 计数/检定符号 → COUNTIF 条件比较符
_COUNTIF_OP = {'=': '=', '<>': '<>', '>': '>', '<': '<',
               '>=': '>=', '<=': '<=', '≠': '<>'}


def script_to_formula_template(replay_cfg: dict, script_name: str) -> str | None:
    """replay_cfg + 脚本名 → 公式模板（{r}/{c} 占位）；不可译返回 None。

    分类（由易到难）：
    - 运算类（operands_raw：column/constant 槽 + 按脚本名操作符）
    - 统计类单区域（range + direction + output；含分位数）
    - 计数（COUNTIF，按行/列分组）
    - 检定（IF+COUNTIF 组合，部分类型）
    其余（查找/排序/任意判定/文本/剪贴板）→ None（用我们机制）。
    """
    raw = replay_cfg.get('operands_raw')
    if raw is not None:
        return _translate_operands(raw, script_name, replay_cfg)
    rng = replay_cfg.get('range')
    if isinstance(rng, (list, tuple)) and len(rng) == 4:
        if replay_cfg.get('inspect_type'):
            return _translate_inspect(replay_cfg)   # 检定（部分类型）
        if 'operator' in replay_cfg:
            return _translate_countif(replay_cfg)   # 计数（按行/列）
        return _translate_stats(rng, replay_cfg, script_name)
    return None


def _translate_operands(raw: list, script_name: str, cfg: dict) -> str | None:
    """运算类：槽位（column/constant）→ 引用片段，按脚本名拼操作符。"""
    if not isinstance(raw, list) or len(raw) < 2:
        return None
    if cfg.get('operands_text'):
        return None   # 文本计算元：MVP 不译
    parts = []
    for s in raw[:2]:
        kind = s.get('kind')
        if kind == 'column':
            parts.append(f"{index_to_col_letter(s['index'])}{{r}}")
        elif kind == 'constant':
            v = s.get('value')
            if v is None:
                return None
            parts.append(_fmt_num(v))
        else:
            return None   # row/剪贴板/文本槽：MVP 不译（干净丢失）
    if len(parts) < 2:
        return None
    if '对数' in script_name:
        # 我们 log(底数, 真数) → Excel LOG(真数, 底数)
        return f'=LOG({parts[1]}, {parts[0]})'
    if '三角' in script_name:
        fn = str(cfg.get('function') or 'sin').lower()
        angle_mode = bool(cfg.get('angle_unit')) \
            and '弧度' not in str(cfg.get('angle_unit'))
        p = parts[0]
        if fn in ('sec', 'csc', 'cot'):
            # Excel 无 sec/csc/cot：拼 1/COS、1/SIN、1/TAN
            inner = f'RADIANS({p})' if angle_mode else p
            return f'=1/{fn.upper().replace("SEC", "COS").replace("CSC", "SIN").replace("COT", "TAN")}({inner})'
        if fn in ('arcsin', 'arccos', 'arctan'):
            # 反三角：Excel 函数名 ASIN/ACOS/ATAN；角度制输出转角度
            inner = f'{fn.replace("arcsin", "ASIN").replace("arccos", "ACOS").replace("arctan", "ATAN")}({p})'
            return f'=DEGREES({inner})' if angle_mode else f'={inner}'
        inner = f'RADIANS({p})' if angle_mode else p
        return f'={fn.upper()}({inner})'
    if '指数' in script_name:
        return f'={parts[0]}^{parts[1]}'
    for k, op in _SCRIPT_BIN_OP.items():
        if k in script_name:
            return f'={parts[0]}{op}{parts[1]}'
    return None


def _translate_stats(rng, cfg: dict, script_name: str) -> str | None:
    """统计类单区域：对列/对行聚合 → 输出行/列（含分位数）。"""
    is_range = '极差' in script_name
    is_quantile = '分位' in script_name
    is_mode = '众数' in script_name
    fn = None
    if not is_range and not is_quantile and not is_mode:
        for k, v in _SCRIPT_AGG_FUNC.items():
            if k in script_name:
                fn = v
                break
    if fn is None and not is_range and not is_quantile and not is_mode:
        return None   # 分位数等：MVP 不译
    r1, c1, r2, c2 = rng
    direction = str(cfg.get('direction', ''))
    out = cfg.get('output') or {}
    if '对列' in direction and out.get('target') == 'row':
        # 每列聚合 → 输出行固定，列跟随（{c}）
        ref = f'{{c}}{r1 + 1}:{{c}}{r2 + 1}'
    elif '对行' in direction and out.get('target') == 'column':
        # 每行聚合 → 输出列固定，行跟随（{r}）
        ref = f'{index_to_col_letter(c1)}{{r}}:{index_to_col_letter(c2)}{{r}}'
    else:
        return None
    if is_range:
        return f'=MAX({ref})-MIN({ref})'
    if is_quantile:
        k = cfg.get('quantile')
        if k is None:
            return None
        try:
            kv = float(k)
        except (TypeError, ValueError):
            return None
        if kv == 0.5:
            return f'=MEDIAN({ref})'   # 中位数（默认）
        return f'=PERCENTILE.INC({ref}, {_fmt_num(kv)})'
    if is_mode:
        # 默认模式 → MODE.SNGL（数值众数，有明确对应）；
        # 精确模式（多众数列表/「无」）→ 无对应，干净丢失（用我们机制）
        if str(cfg.get('mode', '默认')) != '默认':
            return None
        return f'=MODE.SNGL({ref})'
    return f'={fn}({ref})'


def _translate_countif(cfg: dict) -> str | None:
    """计数：按行/列分组 COUNTIF（输出轴垂直，与脚本语义一致）；≡ 不译。"""
    rng = cfg.get('range')
    const = cfg.get('constant')
    if not isinstance(rng, (list, tuple)) or len(rng) != 4 or const is None:
        return None
    op = str(cfg.get('operator', '='))
    if op == '≡':
        return None   # 严格相等（值与写法）：COUNTIF 无法表达
    op = _COUNTIF_OP.get(op, '=')
    ref = _range_ref_with_direction(rng, cfg)
    return f'=COUNTIF({ref}, "{op}{const}")'


def _translate_inspect(cfg: dict) -> str | None:
    """检定：IF(COUNTIF 判定) 组合；任意判定/≡ 不译（用我们机制）。

    检定类型（type）：
    - 存在判定：满足 ≥1 → =IF(COUNTIF(ref,"op常量")>0, pass, fail)
    - 存在型数量：满足 ≥ N → =IF(COUNTIF(ref,"op常量")>=N, pass, fail)
    - 存在型比例：满足/总数 ≥ P → =IF(COUNTIF(ref,"op常量")/COUNTA(ref)>=P, pass, fail)
    - 任意判定：所有数据格都满足 → 无法用 COUNTIF 表达 → None
    """
    rng = cfg.get('range')
    const = cfg.get('constant')
    itype = str(cfg.get('inspect_type', ''))
    if not isinstance(rng, (list, tuple)) or len(rng) != 4 or const is None:
        return None
    op = str(cfg.get('operator', '='))
    if op == '≡' or '任意' in itype:
        return None
    op = _COUNTIF_OP.get(op, '=')
    ref = _range_ref_with_direction(rng, cfg)
    cnt = f'COUNTIF({ref}, "{op}{const}")'
    pass_v = _fmt_literal(cfg.get('pass_result'))
    fail_v = _fmt_literal(cfg.get('fail_result'))
    if pass_v is None or fail_v is None:
        return None
    if '数量' in itype:
        n = cfg.get('type_value')
        if n is None:
            return None
        return f'=IF({cnt}>={_fmt_num(n)}, {pass_v}, {fail_v})'
    if '比例' in itype:
        p = cfg.get('type_value')
        if p is None:
            return None
        return f'=IF({cnt}/COUNTA({ref})>={_fmt_num(p)}, {pass_v}, {fail_v})'
    return f'=IF({cnt}>0, {pass_v}, {fail_v})'   # 存在判定


def _range_ref_with_direction(rng, cfg: dict) -> str:
    """按方向生成区域引用片段（对列→{c} 输出行；对行→{col}{r} 输出列；
    无方向/结构不符 → 整区域绝对引用）。"""
    r1, c1, r2, c2 = rng
    direction = str(cfg.get('direction', ''))
    out = cfg.get('output') or {}
    if '对列' in direction and out.get('target') == 'row':
        return f'{{c}}{r1 + 1}:{{c}}{r2 + 1}'
    if '对行' in direction and out.get('target') == 'column':
        return f'{index_to_col_letter(c1)}{{r}}:{index_to_col_letter(c2)}{{r}}'
    return f'{index_to_col_letter(c1)}{r1 + 1}:{index_to_col_letter(c2)}{r2 + 1}'


def _fmt_literal(v) -> str | None:
    """输出字面量：数字原样、文本加引号；None 返回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return _fmt_num(v)
    return '"' + str(v).replace('"', '""') + '"'


def _fmt_num(v) -> str:
    """数值字面量（整数去 .0）。"""
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else str(f)
    except (TypeError, ValueError):
        return str(v)
