"""公式互译层 — 冒烟测试（动态脚本模式 P1）。

覆盖：
- 引用解析：普通区域/单格/跨sheet/绝对引用/整列/整行/嵌套函数
- 坐标转换：列字母 ↔ 索引
- scan_workbook：真实 xlsx 公式格扫描（含跨 sheet、共享公式展开）
- formula_cell_map 查找表
"""
import os
import sys
import zipfile

sys.path.insert(0, r'D:\Cloude Code\自制表格软件\代码输出库\源码')

from models.formula_translate import (
    parse_formula_refs, parse_cell_ref, parse_range_ref,
    col_letter_to_index, index_to_col_letter, scan_workbook, formula_cell_map,
)
from openpyxl import Workbook

TMP = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_tmp_formula')
os.makedirs(TMP, exist_ok=True)


def check(name, cond):
    if not cond:
        raise AssertionError(f'FAIL: {name}')
    print('PASS:', name)


# ------------------------------------------------------------------
# 1. 坐标转换
# ------------------------------------------------------------------
check('A→0', col_letter_to_index('A') == 0)
check('Z→25', col_letter_to_index('Z') == 25)
check('AA→26', col_letter_to_index('AA') == 26)
check('0→A', index_to_col_letter(0) == 'A')
check('25→Z', index_to_col_letter(25) == 'Z')
check('26→AA', index_to_col_letter(26) == 'AA')
check('parse A1', parse_cell_ref('A1') == (0, 0))
check('parse C10', parse_cell_ref('C10') == (9, 2))
check('parse 非法', parse_cell_ref('X') is None)
check('parse 区域 A1:B3', parse_range_ref('A1:B3') == (0, 0, 2, 1))
check('parse 单格 A1', parse_range_ref('A1') == (0, 0, 0, 0))
check('parse 反区域 B3:A1', parse_range_ref('B3:A1') == (0, 0, 2, 1))

# ------------------------------------------------------------------
# 2. 公式引用解析
# ------------------------------------------------------------------
def refs_plain(formula):
    return [(s, rng) for s, rng in parse_formula_refs(formula)]

check('普通区域 SUM(A1:A3)',
      refs_plain('=SUM(A1:A3)') == [(None, (0, 0, 2, 0))])
check('单格 A1*2', refs_plain('=A1*2') == [(None, (0, 0, 0, 0))])
check('跨sheet 数据表!A1+库存表!A1',
      refs_plain('=数据表!A1+库存表!A1')
      == [('数据表', (0, 0, 0, 0)), ('库存表', (0, 0, 0, 0))])
check('绝对引用 $A$1:$B$3', refs_plain('=$A$1:$B$3') == [(None, (0, 0, 2, 1))])
check('整列 SUM(A:A)', refs_plain('=SUM(A:A)') == [(None, (0, 0, 10**6, 0))])
check('整行 1:1', refs_plain('=SUM(1:1)') == [(None, (0, 0, 0, 10**6))])
check('嵌套 IF(A1>5, B2, C3)',
      refs_plain('=IF(A1>5, B2, C3)')
      == [(None, (0, 0, 0, 0)), (None, (1, 1, 1, 1)), (None, (2, 2, 2, 2))])
check('混合 AVERAGE(Sheet1!A1:C5)+MAX(D1:D2)',
      refs_plain('=AVERAGE(Sheet1!A1:C5)+MAX(D1:D2)')
      == [('Sheet1', (0, 0, 4, 2)), (None, (0, 3, 1, 3))])
check('函数名不误判 SUM(', refs_plain('=SUM(A1)') == [(None, (0, 0, 0, 0))])

# ------------------------------------------------------------------
# 3. scan_workbook 真实文件
# ------------------------------------------------------------------
p = os.path.join(TMP, '公式.xlsx')
wb = Workbook()
ws1 = wb.active
ws1.title = '数据表'
ws1['A1'] = 10
ws1['A2'] = 20
ws1['A3'] = 30
ws1['B1'] = '=SUM(A1:A3)'
ws1['B2'] = '=A1*2'
ws2 = wb.create_sheet('库存表')
ws2['A1'] = 5
ws1['C1'] = '=数据表!A1+库存表!A1'
ws2['B1'] = '=数据表!A1*库存表!A1'
wb.save(p)

cells = scan_workbook(p)
check('扫描到 4 个公式格', len(cells) == 4)
by = {(c['sheet'], tuple(c['cell'])): c for c in cells}
check('B1 公式 SUM', by[('数据表', (0, 1))]['formula'] == '=SUM(A1:A3)')
check('B1 引用解析', [r['range'] for r in by[('数据表', (0, 1))]['refs']] == [[0, 0, 2, 0]])
check('C1 跨sheet引用',
      [r['range'] for r in by[('数据表', (0, 2))]['refs']] == [[0, 0, 0, 0], [0, 0, 0, 0]])
check('C1 引用带来源sheet',
      [r.get('sheet') for r in by[('数据表', (0, 2))]['refs']] == ['数据表', '库存表'])
check('库存表B1 公式', by[('库存表', (0, 1))]['formula'] == '=数据表!A1*库存表!A1')
check('坐标 0-based', (by[('数据表', (1, 1))]['cell']) == [1, 1])

# 共享公式展开：构造一个只有共享公式的独立文件
p2 = os.path.join(TMP, '共享.xlsx')
wb0 = Workbook()
wb0.active['A1'] = 1
wb0.save(p2)
shared_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
 '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
 '<sheetData>'
 '<row r="1"><c r="A1" t="n"><v>10</v></c>'
 '<c r="B1"><f t="shared" ref="B1:B3" si="0">A1*2</f><v>20</v></c></row>'
 '<row r="2"><c r="A2" t="n"><v>20</v></c><c r="B2"><f t="shared" si="0"/><v>40</v></c></row>'
 '<row r="3"><c r="A3" t="n"><v>30</v></c><c r="B3"><f t="shared" si="0"/><v>60</v></c></row>'
 '</sheetData></worksheet>')
z_src = zipfile.ZipFile(p2)
entries = [(item, z_src.read(item)) for item in z_src.namelist()]
z_src.close()
z = zipfile.ZipFile(p2, 'w')
for item, data in entries:
    if item == 'xl/worksheets/sheet1.xml':
        data = shared_xml.encode('utf-8')
    z.writestr(item, data)
z.close()
cells2 = scan_workbook(p2)
b_cells = sorted((c for c in cells2 if c['cell'][1] == 1),
                 key=lambda c: c['cell'][0])
check('共享公式展开为 3 格', len(b_cells) == 3)
check('B2 展开为 A2*2',
      any(c['formula'] == '=A2*2' for c in b_cells))
check('B3 展开为 A3*2',
      any(c['formula'] == '=A3*2' for c in b_cells))

# ------------------------------------------------------------------
# 4. formula_cell_map（含 sheet 维度）
# ------------------------------------------------------------------
fcm = formula_cell_map(cells)
check('查找表含 数据表(0,1)', fcm.get(('数据表', 0, 1)) == '=SUM(A1:A3)')
check('查找表含 库存表(0,1)', fcm.get(('库存表', 0, 1)) == '=数据表!A1*库存表!A1')
check('跨sheet同坐标不互相覆盖', len(fcm) == 4)
check('查找表不含 (9,9)', ('数据表', 9, 9) not in fcm)

# 清理
import shutil
shutil.rmtree(TMP, ignore_errors=True)
print('ALL FORMULA-TRANSLATE TESTS PASSED')
