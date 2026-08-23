"""XLSX 文件读写。使用 openpyxl，MVP 只处理第一个 sheet。"""
from openpyxl import Workbook, load_workbook


def load(file_path: str) -> list[list[str]]:
    """读取 XLSX 文件第一个 sheet，返回二维列表。"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            val = cell.value
            row_data.append(str(val) if val is not None else '')
        rows.append(row_data)
    wb.close()
    return rows


def write(file_path: str, data: list[list[str]]) -> None:
    """写入 XLSX 文件。"""
    wb = Workbook()
    ws = wb.active
    for row_data in data:
        ws.append(row_data)
    wb.save(file_path)
    wb.close()
